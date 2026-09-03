import openpyxl, re
from collections import defaultdict
from openpyxl.utils import column_index_from_string
from map import parse_tab, parse_hdr
S=openpyxl.load_workbook('source.xlsx'); D=openpyxl.load_workbook('payments_by_year.xlsx')
plan=[]; cur=None
for ws in S.worksheets:
    if ws.title in ('Sheet2','Sheet3'): continue
    y1,m1,y2,m2=parse_tab(ws.title)
    if y1 is None: y1=(y2-1) if (y2 is not None and m1>m2) else (y2 if y2 is not None else cur)
    y,m=y1,m1
    for i in range(4):
        plan.append((ws.title,i,y,m)); m+=1
        if m>12: m=1; y+=1
    cur=y
cand=defaultdict(list)
for t,bi,y,m in plan: cand[(y,m)].append((t,bi))
def fill(t,bi):
    w=S[t]; return sum(1 for r in range(1,w.max_row+1) for c in range(bi*3+1,bi*3+4) if w.cell(r,c).value is not None)
primary={k:sorted(v,key=lambda x:-fill(*x))[0] for k,v in cand.items()}
REF=re.compile(r"\$?([A-Z]{1,2})\$?(\d+)")
bad=[];nv=nf=0;ss=ds=0.0;months=0
for (y,m),(t,bi) in sorted(primary.items()):
    sw=S[t]; dw=D[f"{y}年"]; base=(m-1)*3+1; months+=1
    hdr_ok = parse_hdr(sw.cell(1,bi*3+1).value)==(y,m)
    if hdr_ok:
        a=sw.cell(1,bi*3+1); b=dw.cell(1,base)
        if a.value!=b.value or a.font.name!=b.font.name or a.font.sz!=b.font.sz:
            bad.append(('HDR',y,m,1,0,a.value,b.value))
    lo = 2 if hdr_ok else 1
    for r in range(lo, sw.max_row+1):
        for o in range(3):
            a=sw.cell(r,bi*3+1+o).value; b=dw.cell(r+1,base+o).value
            if isinstance(a,str) and a.startswith('='):
                nf+=1
                A=[( (column_index_from_string(c)-1)%3, int(rr)+1) for c,rr in REF.findall(a)]
                B=[( (column_index_from_string(c)-1)%3, int(rr)) for c,rr in REF.findall(b or '')]
                fa=re.sub(r'\$?[A-Z]{1,2}\$?\d+','@',a); fb=re.sub(r'\$?[A-Z]{1,2}\$?\d+','@',b or '')
                if A!=B or fa!=fb: bad.append(('F',y,m,r,o,a,b))
            else:
                nv+=1
                if a!=b: bad.append(('V',y,m,r,o,a,b))
                if isinstance(a,(int,float)): ss+=a
                if isinstance(b,(int,float)): ds+=b
print(f"月ブロック数: {months}  比較セル: 値{nv:,} / 数式{nf:,}  不一致: {len(bad)}")
print(f"数値合計  元={ss:,.0f}  新={ds:,.0f}  差={ss-ds:,.0f}")
for x in bad[:15]: print("  MISMATCH",x)
# every source non-empty cell in A..L accounted for?
tot_src=sum(1 for t,bi,y,m in plan for r in range(1,S[t].max_row+1) for c in range(bi*3+1,bi*3+4) if S[t].cell(r,c).value is not None)
tot_dst=sum(1 for ws in D.worksheets if ws.title.endswith('年') for row in ws.iter_rows() for c in row if c.value is not None)
print(f"元A〜L列の非空セル={tot_src:,}   新ブック全非空セル={tot_dst:,}")
