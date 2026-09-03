import openpyxl, re, json
from copy import copy
from collections import defaultdict
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font, Alignment
from map import parse_tab, parse_hdr, norm

SRC='source.xlsx'; OUT='payments_by_year.xlsx'; SHIFT=1
Z=str.maketrans('0123456789','０１２３４５６７８９')
def wareki(y,m):
    if y>=2019: era,n='令和',y-2018
    elif y>=1989: era,n='平成',y-1988
    else: era,n='昭和',y-1925
    return f"{era}{str(n).translate(Z)}年　{str(m).translate(Z)}月"

wb=openpyxl.load_workbook(SRC)
orig=[ws.title for ws in wb.worksheets]
log=defaultdict(list)

plan=[]; cur=None
for t in orig:
    if t in ('Sheet2','Sheet3'): continue
    ws=wb[t]; y1,m1,y2,m2=parse_tab(t)
    if y1 is None: y1=(y2-1) if (y2 is not None and m1>m2) else (y2 if y2 is not None else cur)
    y,m=y1,m1
    for i in range(4):
        plan.append((t,i,y,m))
        v=ws.cell(1,i*3+1).value; h=parse_hdr(v)
        if h is None:
            log['headers_generated'].append(f"{t} {get_column_letter(i*3+1)}1 → 「{wareki(y,m)}」を新規付与"
                + (f"（元の内容「{norm(v)}」はデータ行として保持）" if v is not None else ""))
        elif h!=(y,m):
            log['headers_corrected'].append(f"{t} {get_column_letter(i*3+1)}1: 元「{norm(v)}」→「{wareki(y,m)}」に訂正")
        m+=1
        if m>12: m=1; y+=1
    cur=y

def fill(t,bi):
    w=wb[t]; return sum(1 for r in range(1,w.max_row+1) for c in range(bi*3+1,bi*3+4) if w.cell(r,c).value is not None)
cand=defaultdict(list)
for t,bi,y,m in plan: cand[(y,m)].append((t,bi))
primary={}; dupes=[]
for k,v in sorted(cand.items()):
    s=sorted(v,key=lambda x:-fill(*x)); primary[k]=s[0]
    for t,bi in s[1:]: dupes.append((k,t,bi))
    if len(v)>1:
        log['overlaps'].append(f"{k[0]}年{k[1]}月 が複数シートに重複: "+
            " / ".join(f"「{t}」第{bi+1}ブロック={fill(t,bi)}セル" for t,bi in s)+
            f" → 「{s[0][0]}」を本体に採用、他は補助領域へ退避")

years=sorted({y for y,_ in primary})
dest={k:(f"{k[0]}年",(k[1]-1)*3+1) for k in primary}
byear={(t,bi):(y,m) for t,bi,y,m in plan}
REF=re.compile(r"(\$?)([A-Z]{1,2})(\$?)(\d+)")

def fx_block(f,t,y):
    def sub(mo):
        d1,col,d2,row=mo.groups(); ci=column_index_from_string(col)
        nrow=int(row)+SHIFT
        if ci>12: return f"{d1}{col}{d2}{nrow}"
        key=byear.get((t,(ci-1)//3))
        if key is None or key not in dest: return mo.group(0)
        ds,base=dest[key]
        if ds!=f"{y}年":
            log['formula_untranslated'].append(f"{t}: {f}"); return mo.group(0)
        return f"{d1}{get_column_letter(base+(ci-1)%3)}{d2}{nrow}"
    return REF.sub(sub,f)

def fx_extra(f,off):
    def sub(mo):
        d1,col,d2,row=mo.groups(); ci=column_index_from_string(col); nrow=int(row)+SHIFT
        return f"{d1}{get_column_letter(ci+off) if ci>12 else col}{d2}{nrow}"
    return REF.sub(sub,f)

def cp(src,dst,rows,c_lo,c_hi,dcol0,fix=None,shift=None):
    for r in rows:
        for c in range(c_lo,c_hi+1):
            sc=src.cell(r,c)
            if sc.value is None and not sc.has_style: continue
            dc=dst.cell(r+(SHIFT if shift is None else shift),dcol0+(c-c_lo)); v=sc.value
            if isinstance(v,str) and v.startswith('=') and fix: v=fix(v)
            dc.value=v
            dc.font=copy(sc.font); dc.fill=copy(sc.fill); dc.border=copy(sc.border)
            dc.alignment=copy(sc.alignment); dc.number_format=sc.number_format
            dc.protection=copy(sc.protection)

for y in years:
    ws=wb.create_sheet(f"{y}年"); contrib={}
    for m in range(1,13):
        if (y,m) not in primary: continue
        t,bi=primary[(y,m)]; src=wb[t]; base=(m-1)*3+1; contrib[t]=src
        for o in range(3):
            sl=get_column_letter(bi*3+1+o)
            if sl in src.column_dimensions and src.column_dimensions[sl].width:
                ws.column_dimensions[get_column_letter(base+o)].width=src.column_dimensions[sl].width
        hdr_ok = parse_hdr(src.cell(1,bi*3+1).value)==(y,m)
        rows = range(2,src.max_row+1) if hdr_ok else range(1,src.max_row+1)
        cp(src,ws,rows,bi*3+1,bi*3+3,base, lambda f,t=t,y=y: fx_block(f,t,y))
        if hdr_ok:
            # 元の見出しセルを書式ごとそのまま1行目へ
            cp(src,ws,[1],bi*3+1,bi*3+1,base,shift=0)
        else:
            c=ws.cell(1,base); c.value=wareki(y,m)
            c.font=Font(name='MS PGothic',size=11,bold=True)
            c.alignment=Alignment(horizontal='center',vertical='center')
    for t,src in contrib.items():
        for r,d in src.row_dimensions.items():
            if d.height:
                k=r+SHIFT; ws.row_dimensions[k].height=max(ws.row_dimensions[k].height or 0,d.height)
    ws.freeze_panes='A2'

placed=defaultdict(lambda:38)
for (y,m),t,bi in dupes:
    ws=wb[f"{y}年"]; src=wb[t]; c0=placed[y]
    l=ws.cell(1,c0); l.value=f"【重複】{t} 第{bi+1}ブロック（{y}年{m}月）"
    l.font=Font(name='MS PGothic',size=9,bold=True,color='FF0000')
    cp(src,ws,range(1,src.max_row+1),bi*3+1,bi*3+3,c0)
    placed[y]=c0+4
for t in orig:
    if t in ('Sheet2','Sheet3'): continue
    src=wb[t]
    if src.max_column<=12: continue
    y=byear[(t,0)][0]; ws=wb[f"{y}年"]; c0=placed[y]; off=c0-13
    l=ws.cell(1,c0); l.value=f"【補助】{t} のM列以降"
    l.font=Font(name='MS PGothic',size=9,bold=True,color='0000FF')
    cp(src,ws,range(1,src.max_row+1),13,src.max_column,c0, lambda f,o=off: fx_extra(f,o))
    for mr in src.merged_cells.ranges:
        if mr.min_col>12:
            ws.merge_cells(start_row=mr.min_row+SHIFT,start_column=mr.min_col+off,
                           end_row=mr.max_row+SHIFT,end_column=mr.max_col+off)
    log['extras'].append(f"{t}: M〜{get_column_letter(src.max_column)}列 → {y}年シート {get_column_letter(c0)}列以降")
    placed[y]=c0+(src.max_column-12)+1

json.dump({k:list(dict.fromkeys(v)) for k,v in log.items()},open('log.json','w'),ensure_ascii=False,indent=1)
wb.save('with_originals.xlsx')
for t in orig:
    if t!='Sheet3': del wb[t]
wb['Sheet3'].title='旧Sheet3（未整理データ）'
wb.save(OUT)
print("sheets:",wb.sheetnames)
for k in ('overlaps','headers_corrected','formula_untranslated'):
    print(f"== {k}: {len(set(log[k]))}")
    for x in dict.fromkeys(log[k]): print("   ",x)
print("headers_generated:",len(log['headers_generated']),"extras:",len(log['extras']))
