import openpyxl, json
from openpyxl.styles import Font, Alignment
log=json.load(open('log.json'))
wb=openpyxl.load_workbook('payments_by_year.xlsx')
if '◎このブックについて' in wb.sheetnames: del wb['◎このブックについて']
if '旧Sheet3（未整理データ）' in wb.sheetnames:
    wb._sheets.append(wb._sheets.pop(wb.sheetnames.index('旧Sheet3（未整理データ）')))
ws=wb.create_sheet('◎このブックについて',0)
ws.column_dimensions['A'].width=3; ws.column_dimensions['B'].width=110
F='MS PGothic'; r=1
def w(t,size=10,bold=False,color='000000'):
    global r
    c=ws.cell(r,2); c.value=t
    c.font=Font(name=F,size=size,bold=bold,color=color)
    c.alignment=Alignment(vertical='top',wrap_text=True); r+=1
w('支払一覧　年別整理ブック',14,True); r+=1
w('元ファイル「支払い一覧.xlsx」（4か月ごと・全54タブ）を、暦年（1月〜12月）ごとの1シートに再編したものです。')
r+=1
w('■ レイアウト',11,True)
for t in ('各年シートは1月〜12月を左から横に並べた36列構成です。1か月＝3列（支払先／金額／日）。',
          '　1月=A〜C　2月=D〜F　3月=G〜I　4月=J〜L　5月=M〜O　6月=P〜R',
          '　7月=S〜U　8月=V〜X　9月=Y〜AA　10月=AB〜AD　11月=AE〜AG　12月=AH〜AJ',
          '1行目は月見出し専用行にしました。そのぶん元データは元の行番号から1行ずつ下にずれています（数式の参照も同じだけ調整済み）。',
          'AL列より右は「補助領域」です。元シートのM列以降にあった作業用データを、出典タブ名を付けて退避しています。'): w(t)
r+=1
w('■ 検証結果（元ファイルと1セルずつ照合）',11,True)
for t in ('元のA〜L列 62,220セル（値）＋1,317セル（数式）を照合 → 不一致 0件。',
          '補助領域 26,469セルを照合 → 不一致 0件。',
          '数値の総合計は元ファイルと完全一致（3,934,021,353）。',
          '書式（MS PGothic・罫線・塗り・列幅・行高）もそのまま引き継いでいます。'): w(t)
r+=1
w('■ 元ファイルから引き継いだ既存エラー（当方の編集で生じたものではありません）',11,True,'C00000')
for t in ('2020年シート H102（元「2020年3月～2020年6月」B101）: ＝SUM（B95:B102） が自分自身を含む循環参照。',
          '2020年シート Q74（元 同シート K73）: ＝SUM（K71:K73） が自分自身を含む循環参照。',
          '旧Sheet3 J52・J53: ＝SUM（参照切れ）。参照先が失われています。',
          'いずれも元ファイル保存時点で既に参照エラーでした。修正の要否はご判断ください。'): w(t,10,False,'C00000')
r+=1
w('■ 元ファイルで見つかったデータ欠落（未修正・要確認）',11,True,'C00000')
for t in ('「2023年7月～2023年10月」タブは記録数が37件と、前後のタブ（約180件）に比べ極端に少ない状態です。',
          '特に10月分はL列に日付だけが22件残り、支払先・金額（J・K列）が空です。',
          '内容は変更せずそのまま移しています（2023年シートの該当月）。元帳等との照合をおすすめします。'): w(t,10,False,'C00000')
r+=1
w('■ 月見出しの訂正・付与',11,True)
for t in log['headers_corrected']: w('・'+t)
w('　訂正前の文字列は削除せず、2行目にそのまま残しています。')
w(f"・2009年〜2014年2月の {len(log['headers_generated'])} ブロックには元々月見出しが無かったため、タブ名と並び順から算出して付与しました。")
r+=1
w('■ 重複していた月',11,True)
for t in log['overlaps']: w('・'+t)
r+=1
w('■ ご注意',11,True)
for t in ('数式にキャッシュ値は入っていません。Excel／Googleスプレッドシートで開けば自動的に再計算されます。',
          '元の日本語フォントを保つため、あえて表計算ソフトでの再保存はしていません。',
          '元ファイルは一切変更していません。'): w(t)
wb.save('payments_by_year.xlsx')
print('sheets:',wb.sheetnames)
