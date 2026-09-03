# -*- coding: utf-8 -*-
"""ETC利用明細PDFを解析し、車番（＝担当ドライバー）ごとに通行料金を振り分ける。

代車・乗り替わりがあった日は OVERRIDES に登録すると、その日の明細だけを
別の車番（担当者）へ付け替える。
"""
import re, collections, datetime, pdfplumber, openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

PDF  = '/root/.claude/uploads/e46d2994-6f12-5b86-9532-97b1f22855c3/6be19bcb-____.pdf'
XLSX = '/root/.claude/uploads/e46d2994-6f12-5b86-9532-97b1f22855c3/e7bb02bd-____.xlsx'
OUT  = '/home/user/nisshou/8月分_高速料金_車番別.xlsx'

# 乗り替わりの付け替え設定
#   (実際に走った車番, 利用日(YY/MM/DD)) -> 料金を負担する車番（担当者の車番）
OVERRIDES = {
    (46, '26/08/20'): 9283,   # 8/20・8/21は車番46に伊垣さん(9283)が乗車
    (46, '26/08/21'): 9283,
}

date_re = re.compile(r'^(\d{2}/\d{2}/\d{2})(?:\s+(\d{2}/\d{2}/\d{2}))?\b')
time_re = re.compile(r'^\d{2}:\d{2}(\s+\d{2}:\d{2})?\b')
card_re = re.compile(r'\*{3,}\d+$')


def parse(path):
    """明細PDFを1レコード=3行として解析する。

    利用年月日(自) 利用年月日(至) (割引前料金) 車種   備考
    時分(自)       時分(至)       (ETC割引額)  車両番号
    利用IC(自)     利用IC(至)     通行料金     ETCカード番号

    本線料金所の単独課金は日付・時刻が1つしか出力されないため、
    日付行・時刻行はいずれも「1つまたは2つ」を許容する。
    """
    lines = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            lines += (page.extract_text() or '').split('\n')

    recs, bad = [], []
    d1 = d2 = t1 = t2 = car = None
    note = ''
    for ln in lines:
        s = ln.strip()
        m = date_re.match(s)
        if m:
            d1, d2 = m.group(1), m.group(2) or m.group(1)
            last = s.split()[-1]
            note = '' if re.fullmatch(r'\d+', last) else last
            continue
        m = time_re.match(s)
        if m:
            toks = s.split()
            t1, t2 = toks[0], (toks[1] if m.group(1) else toks[0])
            car = int(toks[-1]) if re.fullmatch(r'\d+', toks[-1]) else None
            if car is None:
                bad.append(s)
            continue
        if card_re.search(s):
            toks = s.split()
            fee = toks[-2].replace(',', '')
            if car is None or not re.fullmatch(r'\d+', fee):
                bad.append(s)
                continue
            recs.append(dict(car=car, d1=d1, d2=d2, t1=t1, t2=t2,
                             fee=int(fee), ic=' '.join(toks[:-2]), note=note))
            car = None
    if bad:
        raise SystemExit('解析できない行があります: %r' % bad)
    return recs


def to_date(s):
    """'26/08/01' -> date(2026, 8, 1)"""
    y, m, d = (int(x) for x in s.split('/'))
    return datetime.date(2000 + y, m, d)


def build_detail_sheet(wb, recs, names, order):
    """車番別の利用明細シートを作る。"""
    if '車番別明細' in wb.sheetnames:
        del wb['車番別明細']
    ws = wb.create_sheet('車番別明細')
    YEN = '"￥"#,##0'
    head = ['車番', '担当', '利用日', '時分', '到着日', '時分',
            '入口ＩＣ', '出口ＩＣ', '通行料金', '割引', '備考']
    thin = Side(style='thin', color='BFBFBF')
    for i, h in enumerate(head, 1):
        c = ws.cell(1, i, h)
        c.font = Font(name='游ゴシック', sz=11, b=True)
        c.fill = PatternFill('solid', fgColor='DDEBF7')
        c.alignment = Alignment(horizontal='center')
        c.border = Border(bottom=thin)

    by_car = collections.defaultdict(list)
    for r in recs:
        by_car[r['charge_to']].append(r)

    row = 2
    for car in order:
        rs = sorted(by_car.get(car, []), key=lambda r: (r['d1'], r['t1']))
        if not rs:
            continue
        for r in rs:
            same_day = r['d2'] == r['d1']
            ic = r['ic'].split()
            note = []
            if len(ic) == 1:
                ic = [ic[0], '']
                note.append('単独課金')
            if r['car'] != car:
                note.append('実車番%s に乗車' % r['car'])
            vals = [car, names.get(car, ''), to_date(r['d1']), r['t1'],
                    None if same_day else to_date(r['d2']), r['t2'],
                    ic[0], ic[1], r['fee'], r['note'], '／'.join(note) or None]
            for i, v in enumerate(vals, 1):
                c = ws.cell(row, i, v)
                c.font = Font(name='游ゴシック', sz=10)
            ws.cell(row, 3).number_format = 'yyyy/m/d'
            ws.cell(row, 5).number_format = 'yyyy/m/d'
            ws.cell(row, 9).number_format = YEN
            row += 1
        c = ws.cell(row, 8, '小計')
        c.alignment = Alignment(horizontal='right')
        ws.cell(row, 9, sum(r['fee'] for r in rs)).number_format = YEN
        for i in range(1, 12):
            cell = ws.cell(row, i)
            cell.font = Font(name='游ゴシック', sz=10, b=True)
            cell.border = Border(top=thin, bottom=thin)
        row += 2

    ws.cell(row, 8, '総合計').font = Font(name='游ゴシック', sz=11, b=True)
    ws.cell(row, 8).alignment = Alignment(horizontal='right')
    ws.cell(row, 9, sum(r['fee'] for r in recs)).number_format = YEN
    ws.cell(row, 9).font = Font(name='游ゴシック', sz=11, b=True)

    for col, w in zip('ABCDEFGHIJK', (7, 12, 11, 7, 11, 7, 16, 16, 11, 13, 20)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = 'A2'
    return ws


def main():
    recs = parse(PDF)
    grand = sum(r['fee'] for r in recs)
    print('明細件数: %d  通行料金合計: ￥%s' % (len(recs), format(grand, ',')))

    totals, counts = collections.Counter(), collections.Counter()
    moved = collections.Counter()
    for r in recs:
        to = OVERRIDES.get((r['car'], r['d1']), r['car'])
        r['charge_to'] = to
        if to != r['car']:
            moved[(r['car'], to)] += r['fee']
        totals[to] += r['fee']
        counts[to] += 1
    for (frm, to), amt in sorted(moved.items()):
        print('付け替え: 車番%s → 車番%s  ￥%s' % (frm, to, format(amt, ',')))
    assert sum(totals.values()) == grand

    wb = openpyxl.load_workbook(XLSX)
    ws = wb['高速料金']
    YEN = '"￥"#,##0'
    listed = set()
    for row in range(2, ws.max_row + 1):
        car = ws.cell(row, 1).value
        if isinstance(car, int):
            listed.add(car)
            ws.cell(row, 3).value = totals.get(car, 0)
            ws.cell(row, 3).number_format = YEN
        else:
            ws.cell(row, 3).value = None

    extra = sorted(set(totals) - listed)
    row = ws.max_row + 1
    order = [ws.cell(r, 1).value for r in range(2, row)
             if isinstance(ws.cell(r, 1).value, int)] + extra
    for car in extra:
        ws.cell(row, 1).value = car
        ws.cell(row, 2).value = '（一覧に無い車番）'
        ws.cell(row, 3).value = totals[car]
        ws.cell(row, 3).number_format = YEN
        row += 1

    ws['D1'] = '備考'
    ws['D1'].font = Font(name='游ゴシック', sz=11, b=True)

    # 付け替えのあった行に備考を入れる
    notes = collections.defaultdict(list)
    for (frm, to), amt in moved.items():
        days = sorted(d for (c, d) in OVERRIDES if c == frm and OVERRIDES[(c, d)] == to)
        days = '・'.join('%d/%d' % (int(d.split('/')[1]), int(d.split('/')[2])) for d in days)
        notes[frm].append('%s分 ￥%s を車番%s へ振替' % (days, format(amt, ','), to))
        notes[to].append('車番%s の %s分 ￥%s を含む' % (frm, days, format(amt, ',')))
    for r in range(2, row):
        car = ws.cell(r, 1).value
        if isinstance(car, int) and car in notes:
            ws.cell(r, 4).value = '／'.join(notes[car])
            ws.cell(r, 4).font = Font(name='游ゴシック', sz=10)

    tr = row + 1
    ws.cell(tr, 2).value = '合計'
    ws.cell(tr, 3).value = sum(totals.values())
    ws.cell(tr, 3).number_format = YEN
    for col in (2, 3):
        ws.cell(tr, col).font = Font(name='游ゴシック', sz=12, b=True)
    for r in range(2, tr + 1):
        for col in (1, 2, 3):
            cell = ws.cell(r, col)
            if cell.font.name is None:
                cell.font = Font(name='游ゴシック', sz=12)
    ws.column_dimensions['C'].width = 13
    ws.column_dimensions['D'].width = 34

    names = {}
    for r in range(2, tr):
        car, nm = ws.cell(r, 1).value, ws.cell(r, 2).value
        if isinstance(car, int) and nm:
            names[car] = nm
    build_detail_sheet(wb, recs, names, order)
    wb.save(OUT)

    print('明細に出現しない車番:', sorted(c for c in listed if c not in totals))
    print('一覧に無い車番:', extra)
    print('saved', OUT)


if __name__ == '__main__':
    main()
